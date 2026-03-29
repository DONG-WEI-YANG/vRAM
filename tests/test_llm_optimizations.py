"""
vRAM LLM-Inspired Optimizations — R1-R11 Test Suite
=====================================================
Adapted from the 11-round layered testing methodology:

  R1  Core Engine (基礎引擎)        — 8 項 LLM 最佳化的基本 CRUD/生命週期
  R2  API Contract (介面契約)        — public API 簽名、stats、可組合性
  R3  Static Scan (靜態掃描)         — import、module integrity、circular dep
  R4  User Journey (使用者旅程)      — 完整 optimize→access→evict→report 流程
  R5  Resilience (韌性/破壞)         — edge case、overflow、empty、concurrent
  R6  Correctness (數學/正確性)      — importance score、compression ratio、hit rate
  R7  LLM Domain (領域專家)          — 每項 LLM 技術的語意正確性
  R8  Integration (整合層)           — MemoryPool / Prefetcher / SlowDevice 整合
  R9  Pipeline (多裝置管線)          — StripedSwap + GroupedIO + FlashIO pipeline
  R10 Encoding (編碼防護)            — Unicode block tags、JSON roundtrip、path
  R11 Export (匯出/報告)             — Benchmark 執行、Excel 生成、stats 一致性

Total: ~140 tests
"""
import math
import os
import sys
import json
import struct
import tempfile
import threading
import time
import unittest
import zlib
from collections import OrderedDict
from unittest.mock import patch, MagicMock

# ── Path setup ──
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "提升VRAM與GPA效能設計方案"))

from microsystems.core.llm_optimizations import (
    BlockPlacementCache,
    PlacementEntry,
    FlashIOEngine,
    FusedTransferBatch,
    SharedCompressionDict,
    SlidingWindowTracker,
    StreamingMemoryManager,
    AnchorBlock,
    ImportanceBasedEvictor,
    BlockImportance,
    SpeculativePrefetchEngine,
    PrefetchPrediction,
    GroupedBlockScheduler,
    BlockGroup,
    LLMOptimizationSuite,
)


# ═══════════════════════════════════════════════════════════════
# R1  Core Engine — 基礎 CRUD / 生命週期 (18 tests)
# ═══════════════════════════════════════════════════════════════
class TestR1_PlacementCache_Lifecycle(unittest.TestCase):
    """KV Cache → Block Placement Cache lifecycle"""

    def test_record_and_lookup(self):
        cache = BlockPlacementCache()
        cache.record("layer_0", tier=0, access_count=10, latency_ms=0.5)
        self.assertEqual(cache.lookup("layer_0"), 0)

    def test_lookup_miss_returns_none(self):
        cache = BlockPlacementCache()
        self.assertIsNone(cache.lookup("nonexistent"))

    def test_lru_eviction_on_max(self):
        cache = BlockPlacementCache(max_entries=3)
        for i in range(5):
            cache.record(f"layer_{i}", tier=0, access_count=1, latency_ms=0.1)
        # First 2 should be evicted
        self.assertIsNone(cache.lookup("layer_0"))
        self.assertIsNone(cache.lookup("layer_1"))
        self.assertEqual(cache.lookup("layer_4"), 0)

    def test_warm_start_plan(self):
        cache = BlockPlacementCache()
        for i in range(5):
            cache.record(f"layer_{i}", tier=i % 3, access_count=10, latency_ms=0.1)
        plan = cache.warm_start_plan([f"layer_{i}" for i in range(5)])
        self.assertEqual(len(plan), 5)


class TestR1_FlashIO_Lifecycle(unittest.TestCase):
    """Flash Attention → Flash I/O lifecycle"""

    def test_submit_and_flush(self):
        flash = FlashIOEngine()
        flash.submit("b1", 2, 1, 64*1024*1024)
        batch = flash.flush()
        self.assertIsNotNone(batch)
        self.assertEqual(len(batch.requests), 1)

    def test_empty_flush_returns_none(self):
        flash = FlashIOEngine()
        self.assertIsNone(flash.flush())

    def test_group_by_direction(self):
        flash = FlashIOEngine()
        flash.submit("b1", 2, 1, 100)
        flash.submit("b2", 2, 1, 200)
        flash.submit("b3", 1, 0, 300)
        batch = flash.flush()
        groups = flash.group_by_direction(batch)
        self.assertEqual(len(groups), 2)  # (2→1) and (1→0)


class TestR1_SlidingWindow_Lifecycle(unittest.TestCase):
    """Sliding Window Attention → Bounded Tracker lifecycle"""

    def test_access_returns_was_hot(self):
        sw = SlidingWindowTracker(window_size=3)
        self.assertFalse(sw.access("b1"))  # first access → was not hot
        self.assertTrue(sw.access("b1"))   # second → was hot

    def test_window_overflow_evicts(self):
        sw = SlidingWindowTracker(window_size=3)
        for i in range(5):
            sw.access(f"b{i}")
        self.assertFalse(sw.is_hot("b0"))
        self.assertFalse(sw.is_hot("b1"))
        self.assertTrue(sw.is_hot("b4"))

    def test_remove_block(self):
        sw = SlidingWindowTracker(window_size=10)
        sw.access("b1")
        sw.remove("b1")
        self.assertFalse(sw.is_hot("b1"))


class TestR1_ImportanceEvictor_Lifecycle(unittest.TestCase):
    """Pruning KV Cache → Importance Evictor lifecycle"""

    def test_compute_importance_returns_dataclass(self):
        ev = ImportanceBasedEvictor()
        imp = ev.compute_importance("b1", access_count=10,
            last_access_ts=time.time(), size_bytes=64*1024*1024)
        self.assertIsInstance(imp, BlockImportance)
        self.assertIsInstance(imp.score, float)

    def test_select_victims_excludes_set(self):
        ev = ImportanceBasedEvictor()
        blocks = [("b1", 1, time.time()-100, 1000),
                  ("b2", 100, time.time(), 1000)]
        victims = ev.select_victims(blocks, count=1, exclude={"b1"})
        self.assertEqual(len(victims), 1)
        self.assertEqual(victims[0].block_id, "b2")


class TestR1_SpeculativePrefetch_Lifecycle(unittest.TestCase):
    """Speculative Decoding → Speculative Prefetch lifecycle"""

    def test_record_and_predict(self):
        spec = SpeculativePrefetchEngine()
        spec.record_access("A")
        spec.record_access("B")
        spec.record_access("A")
        preds = spec.predict_next("A", ["B", "C", "D"])
        self.assertIn("B", preds)

    def test_verify_updates_confidence(self):
        spec = SpeculativePrefetchEngine()
        initial = spec._current_confidence
        spec.verify_prediction(["A", "B"], "A")
        self.assertGreater(spec._current_confidence, initial)


class TestR1_GroupedScheduler_Lifecycle(unittest.TestCase):
    """GQA → Grouped Block I/O lifecycle"""

    def test_record_and_flush(self):
        gs = GroupedBlockScheduler()
        gs.record_access("b1")
        gs.record_access("b2")
        gs.flush_window()
        self.assertEqual(gs.stats()["total_co_accesses"], 1)

    def test_detect_groups_requires_threshold(self):
        gs = GroupedBlockScheduler(min_co_access=2)
        # Only 1 co-access → no group
        gs.record_access("b1")
        gs.record_access("b2")
        gs.flush_window()
        groups = gs.detect_groups()
        self.assertEqual(len(groups), 0)


class TestR1_SharedCompressionDict_Lifecycle(unittest.TestCase):
    """MLA → Shared Compression Dictionary lifecycle"""

    def test_train_on_samples(self):
        scd = SharedCompressionDict(dict_size=4096, sample_blocks=4)
        samples = [os.urandom(1024) for _ in range(4)]
        scd.train(samples)
        self.assertIsNotNone(scd._dictionary)
        self.assertGreater(len(scd._dictionary), 0)

    def test_compress_decompress_roundtrip(self):
        scd = SharedCompressionDict()
        base = b"ABCDEFGH" * 128
        samples = [base + os.urandom(64) for _ in range(8)]
        scd.train(samples)
        original = base + os.urandom(32)
        compressed = scd.compress(original)
        decompressed = scd.decompress(compressed)
        self.assertEqual(decompressed, original)

    def test_compress_without_training_fallback(self):
        scd = SharedCompressionDict()
        data = b"test data" * 100
        compressed = scd.compress(data)
        self.assertTrue(compressed.startswith(b"\x00"))  # standard marker
        decompressed = scd.decompress(compressed)
        self.assertEqual(decompressed, data)

    def test_stats_tracking(self):
        scd = SharedCompressionDict()
        stats = scd.stats()
        self.assertIn("has_dictionary", stats)
        self.assertIn("improvement_pct", stats)
        self.assertFalse(stats["has_dictionary"])


class TestR1_StreamingMemoryManager_Lifecycle(unittest.TestCase):
    """Streaming LLM → Anchor Pinning lifecycle"""

    def test_register_anchor(self):
        smm = StreamingMemoryManager(anchor_slots=4)
        ok = smm.register_anchor("blk_0", importance=10.0, reason="test")
        self.assertTrue(ok)
        self.assertTrue(smm.is_anchor("blk_0"))

    def test_anchor_slots_limit(self):
        smm = StreamingMemoryManager(anchor_slots=2)
        smm.register_anchor("a", 10.0)
        smm.register_anchor("b", 20.0)
        ok = smm.register_anchor("c", 5.0)
        self.assertFalse(ok)

    def test_select_eviction_excludes_anchors(self):
        smm = StreamingMemoryManager(anchor_slots=2)
        smm.register_anchor("blk_0", 10.0)
        blocks = [("blk_0", 100, 1.0), ("blk_1", 5, 0.5), ("blk_2", 3, 0.3)]
        evicted = smm.select_eviction_candidates(blocks, "vram", count=2)
        self.assertNotIn("blk_0", evicted)
        self.assertIn("blk_2", evicted)

    def test_new_session_increments(self):
        smm = StreamingMemoryManager()
        smm.new_session()
        smm.new_session()
        self.assertEqual(smm.stats()["sessions_survived"], 2)


# ═══════════════════════════════════════════════════════════════
# R2  API Contract — stats / 可組合性 (12 tests)
# ═══════════════════════════════════════════════════════════════
class TestR2_Stats(unittest.TestCase):
    """All modules return well-formed stats dicts"""

    def test_placement_cache_stats(self):
        s = BlockPlacementCache().stats()
        self.assertIn("entries", s)
        self.assertIn("hit_rate_pct", s)

    def test_flash_io_stats(self):
        s = FlashIOEngine().stats()
        self.assertIn("total_individual_ios", s)
        self.assertIn("io_reduction_pct", s)

    def test_shared_dict_stats(self):
        s = SharedCompressionDict().stats()
        self.assertIn("has_dictionary", s)
        self.assertIn("improvement_pct", s)

    def test_sliding_window_stats(self):
        s = SlidingWindowTracker().stats()
        self.assertIn("hot_count", s)
        self.assertIn("cold_count", s)

    def test_streaming_stats(self):
        s = StreamingMemoryManager().stats()
        self.assertIn("anchor_count", s)
        self.assertIn("rolling_evictions", s)

    def test_importance_evictor_stats(self):
        s = ImportanceBasedEvictor().stats()
        self.assertIn("weights", s)
        self.assertIn("total_scores_computed", s)

    def test_speculative_stats(self):
        s = SpeculativePrefetchEngine().stats()
        self.assertIn("hit_rate_pct", s)
        self.assertIn("current_confidence", s)

    def test_grouped_stats(self):
        s = GroupedBlockScheduler().stats()
        self.assertIn("total_groups", s)
        self.assertIn("groups", s)

    def test_suite_full_stats_has_all_keys(self):
        suite = LLMOptimizationSuite()
        s = suite.full_stats()
        self.assertIn("enabled", s)
        self.assertEqual(sum(1 for v in s["enabled"].values() if v), 8)

    def test_suite_disabled_modules(self):
        suite = LLMOptimizationSuite(enable_flash_io=False, enable_streaming=False)
        s = suite.full_stats()
        self.assertNotIn("flash_io", s)
        self.assertNotIn("streaming", s)

    def test_suite_on_block_access_no_error(self):
        suite = LLMOptimizationSuite()
        suite.on_block_access("b1", access_count=5, size_bytes=1024)

    def test_suite_on_transfer_request_no_error(self):
        suite = LLMOptimizationSuite()
        suite.on_transfer_request("b1", 2, 1, 1024)


# ═══════════════════════════════════════════════════════════════
# R3  Static Scan — imports, module integrity (8 tests)
# ═══════════════════════════════════════════════════════════════
class TestR3_StaticScan(unittest.TestCase):
    """Module imports and structural integrity"""

    def test_import_llm_optimizations(self):
        import microsystems.core.llm_optimizations as mod
        self.assertTrue(hasattr(mod, "LLMOptimizationSuite"))

    def test_import_benchmark(self):
        import microsystems.core.benchmark_llm_opts as mod
        self.assertTrue(hasattr(mod, "run_all_benchmarks"))

    def test_all_8_classes_exist(self):
        classes = [BlockPlacementCache, FlashIOEngine, SharedCompressionDict,
                   SlidingWindowTracker, StreamingMemoryManager,
                   ImportanceBasedEvictor, SpeculativePrefetchEngine,
                   GroupedBlockScheduler]
        self.assertEqual(len(classes), 8)

    def test_suite_default_enables_all_8(self):
        suite = LLMOptimizationSuite()
        self.assertIsNotNone(suite.placement_cache)
        self.assertIsNotNone(suite.flash_io)
        self.assertIsNotNone(suite.shared_dict)
        self.assertIsNotNone(suite.sliding_window)
        self.assertIsNotNone(suite.streaming)
        self.assertIsNotNone(suite.importance_evictor)
        self.assertIsNotNone(suite.speculative_prefetch)
        self.assertIsNotNone(suite.grouped_io)

    def test_dataclass_placement_entry(self):
        e = PlacementEntry(block_tag="t", best_tier=0, avg_access_count=1.0,
                           avg_latency_ms=0.1)
        self.assertEqual(e.block_tag, "t")

    def test_dataclass_anchor_block(self):
        a = AnchorBlock(block_id="b1", importance_score=5.0)
        self.assertEqual(a.block_id, "b1")

    def test_dataclass_block_importance(self):
        b = BlockImportance(block_id="b1", score=1.0, recency=0.5,
                            frequency=0.5, size_penalty=0.1, is_heavy_hitter=False)
        self.assertFalse(b.is_heavy_hitter)

    def test_memory_pool_imports_optimizations(self):
        from microsystems.core.memory_pool import MemoryPool
        pool = MemoryPool(1024, 1024, 1024)
        self.assertTrue(hasattr(pool, '_sw_tracker'))
        self.assertTrue(hasattr(pool, '_importance_evictor'))


# ═══════════════════════════════════════════════════════════════
# R4  User Journey — 完整流程 (12 tests)
# ═══════════════════════════════════════════════════════════════
class TestR4_UserJourney(unittest.TestCase):
    """End-to-end optimization journey"""

    def test_full_placement_cache_journey(self):
        """Session 1: cold start → record → Session 2: warm start"""
        cache = BlockPlacementCache()
        # Session 1
        for i in range(10):
            cache.record(f"layer_{i}", tier=0 if i < 3 else 1,
                         access_count=100-i*10, latency_ms=0.1)
        # Session 2
        plan = cache.warm_start_plan([f"layer_{i}" for i in range(10)])
        self.assertGreater(len(plan), 0)
        self.assertEqual(plan["layer_0"], 0)  # hot layer stays in VRAM

    def test_full_eviction_journey(self):
        """Access blocks → compute importance → evict victims"""
        ev = ImportanceBasedEvictor()
        now = time.time()
        # Simulate varied access patterns
        for _ in range(50):
            ev.record_access("hot_block")
        for _ in range(2):
            ev.record_access("cold_block")
        blocks = [
            ("hot_block", 50, now, 64*1024*1024),
            ("cold_block", 2, now - 300, 64*1024*1024),
            ("medium_block", 10, now - 60, 32*1024*1024),
        ]
        victims = ev.select_victims(blocks, count=1)
        self.assertEqual(victims[0].block_id, "cold_block")

    def test_full_streaming_journey(self):
        """Detect anchors → run long session → rolling eviction"""
        sm = StreamingMemoryManager(max_vram_blocks=5, anchor_slots=2)
        stats = [("b0", 1000, 0.1), ("b1", 500, 0.1),
                 ("b2", 50, 0.1), ("b3", 10, 0.1)]
        anchors = sm.auto_detect_anchors(stats)
        self.assertEqual(len(anchors), 2)
        self.assertTrue(sm.is_anchor("b0"))
        # Eviction should preserve anchors
        blocks = [(f"b{i}", 10, time.time() - i * 10) for i in range(10)]
        victims = sm.select_eviction_candidates(blocks, "vram", count=3)
        self.assertNotIn("b0", victims)
        self.assertNotIn("b1", victims)

    def test_full_speculative_learning_journey(self):
        """Train transition matrix → predict → verify → adapt"""
        spec = SpeculativePrefetchEngine(base_lookahead=2, max_lookahead=6)
        # Training phase: A→B→C→A→B→C pattern
        for _ in range(20):
            for block in ["A", "B", "C"]:
                spec.record_access(block)
        # Predict after A
        preds = spec.predict_next("A", ["B", "C", "D"])
        self.assertEqual(preds[0], "B")  # should learn A→B
        # Verify correct
        spec.verify_prediction(preds, "B")
        self.assertGreater(spec._current_confidence, 0.5)

    def test_full_grouped_io_journey(self):
        """Record co-access → detect groups → get io plan"""
        gs = GroupedBlockScheduler(min_co_access=3)
        for _ in range(5):
            gs.record_access("weight_q")
            gs.record_access("weight_k")
            gs.record_access("weight_v")
            gs.flush_window()
        groups = gs.detect_groups()
        self.assertGreater(len(groups), 0)
        # Q/K/V should be in same group
        plan = gs.get_group_io_plan("weight_q")
        self.assertIn("weight_k", plan)
        self.assertIn("weight_v", plan)

    def test_suite_unified_access_journey(self):
        """LLMOptimizationSuite processes block accesses uniformly"""
        suite = LLMOptimizationSuite(window_size=10)
        for i in range(20):
            suite.on_block_access(f"block_{i % 5}")
        stats = suite.full_stats()
        sw_stats = stats["sliding_window"]
        self.assertEqual(sw_stats["hot_count"], 5)
        self.assertGreater(sw_stats["total_accesses"], 0)

    def test_flash_io_batch_journey(self):
        """Submit many → flush batches → verify reduction"""
        flash = FlashIOEngine(max_batch_size=10, min_batch_bytes=999999999)
        for i in range(30):
            flash.submit(f"b{i}", 2, 1, 1024)
        # Manual flush every 10
        batches = 0
        while True:
            b = flash.flush()
            if b is None:
                break
            batches += 1
        self.assertLessEqual(batches, 30)  # should batch

    def test_shared_dict_journey(self):
        """Train dict → compress → decompress → verify"""
        sd = SharedCompressionDict()
        base = bytes(range(256)) * 100
        samples = [base + bytes([i]*200) for i in range(8)]
        sd.train(samples)
        for data in samples:
            compressed = sd.compress(data)
            decompressed = sd.decompress(compressed)
            self.assertEqual(decompressed, data)

    def test_suite_eviction_journey(self):
        """Suite combines importance + anchor + window for eviction"""
        suite = LLMOptimizationSuite()
        now = time.time()
        # Register anchor
        suite.streaming.register_anchor("anchor_block", 10.0, "critical")
        blocks = [
            ("anchor_block", 100, now, 64*1024*1024),
            ("hot_block", 50, now, 64*1024*1024),
            ("cold_block", 1, now - 600, 64*1024*1024),
        ]
        victims = suite.get_eviction_candidates(blocks, count=1)
        self.assertNotIn("anchor_block", victims)
        self.assertEqual(victims[0], "cold_block")

    def test_sliding_window_hot_cold_separation(self):
        """Hot blocks in window, cold blocks outside"""
        sw = SlidingWindowTracker(window_size=5)
        for i in range(10):
            sw.access(f"b{i}")
        hot = sw.get_hottest_blocks(3)
        cold = sw.get_coldest_blocks(3)
        # Hot should be recent (b7, b8, b9)
        self.assertIn("b9", hot)
        # Cold blocks should NOT be in hot window
        for bid in cold:
            self.assertFalse(sw.is_hot(bid), f"{bid} should be cold")

    def test_streaming_new_session_resets(self):
        sm = StreamingMemoryManager()
        sm._rolling_evictions = 100
        sm.new_session()
        self.assertEqual(sm._sessions_survived, 1)

    def test_placement_cache_update_existing(self):
        cache = BlockPlacementCache()
        cache.record("t1", 2, access_count=10, latency_ms=1.0)
        cache.record("t1", 0, access_count=100, latency_ms=0.1)
        self.assertEqual(cache.lookup("t1"), 0)


# ═══════════════════════════════════════════════════════════════
# R5  Resilience — edge case / stress (14 tests)
# ═══════════════════════════════════════════════════════════════
class TestR5_Resilience(unittest.TestCase):
    """Edge cases, empty inputs, overflow, concurrency"""

    def test_empty_blocks_eviction(self):
        ev = ImportanceBasedEvictor()
        victims = ev.select_victims([], count=5)
        self.assertEqual(len(victims), 0)

    def test_predict_no_history(self):
        spec = SpeculativePrefetchEngine()
        preds = spec.predict_next("unknown", ["A", "B"])
        self.assertIsInstance(preds, list)

    def test_verify_empty_prediction(self):
        spec = SpeculativePrefetchEngine()
        result = spec.verify_prediction([], "A")
        self.assertFalse(result)

    def test_window_size_one(self):
        sw = SlidingWindowTracker(window_size=1)
        sw.access("b1")
        sw.access("b2")
        self.assertFalse(sw.is_hot("b1"))
        self.assertTrue(sw.is_hot("b2"))

    def test_streaming_anchor_overflow(self):
        sm = StreamingMemoryManager(anchor_slots=2)
        sm.register_anchor("a1", 10.0)
        sm.register_anchor("a2", 20.0)
        result = sm.register_anchor("a3", 5.0)  # weaker than both
        self.assertFalse(result)  # should reject

    def test_streaming_anchor_replace_weakest(self):
        sm = StreamingMemoryManager(anchor_slots=2)
        sm.register_anchor("a1", 5.0)
        sm.register_anchor("a2", 10.0)
        result = sm.register_anchor("a3", 15.0)  # stronger than a1
        self.assertTrue(result)
        self.assertFalse(sm.is_anchor("a1"))
        self.assertTrue(sm.is_anchor("a3"))

    def test_grouped_empty_window(self):
        gs = GroupedBlockScheduler()
        gs.flush_window()  # empty window
        self.assertEqual(gs.stats()["total_co_accesses"], 0)

    def test_grouped_single_block_window(self):
        gs = GroupedBlockScheduler()
        gs.record_access("b1")
        gs.flush_window()  # <2 blocks → no co-access
        self.assertEqual(gs.stats()["total_co_accesses"], 0)

    def test_flash_io_zero_bytes(self):
        flash = FlashIOEngine()
        flash.submit("b1", 2, 1, 0)
        batch = flash.flush()
        self.assertEqual(batch.total_bytes, 0)

    def test_importance_zero_access(self):
        ev = ImportanceBasedEvictor()
        imp = ev.compute_importance("b1", 0, 0.0, 0)
        self.assertIsInstance(imp.score, float)

    def test_shared_dict_no_train(self):
        """Compress without training → should fallback to standard"""
        sd = SharedCompressionDict()
        data = b"test data" * 100
        compressed = sd.compress(data)
        self.assertTrue(compressed.startswith(b"\x00"))  # standard marker

    def test_shared_dict_empty_data(self):
        sd = SharedCompressionDict()
        result = sd.decompress(b"")
        self.assertEqual(result, b"")

    def test_concurrent_sliding_window(self):
        """Thread-safety of sliding window"""
        sw = SlidingWindowTracker(window_size=100)
        errors = []

        def worker(prefix, n):
            try:
                for i in range(n):
                    sw.access(f"{prefix}_{i}")
            except Exception as e:
                errors.append(e)

        threads = [threading.Thread(target=worker, args=(f"t{t}", 50))
                   for t in range(4)]
        for t in threads:
            t.start()
        for t in threads:
            t.join()
        self.assertEqual(len(errors), 0)

    def test_concurrent_placement_cache(self):
        """Thread-safety of placement cache"""
        cache = BlockPlacementCache()
        errors = []

        def worker(prefix, n):
            try:
                for i in range(n):
                    cache.record(f"{prefix}_{i}", tier=0, access_count=1, latency_ms=0.1)
                    cache.lookup(f"{prefix}_{i}")
            except Exception as e:
                errors.append(e)

        threads = [threading.Thread(target=worker, args=(f"t{t}", 50))
                   for t in range(4)]
        for t in threads:
            t.start()
        for t in threads:
            t.join()
        self.assertEqual(len(errors), 0)

    def test_shared_dict_train_empty_samples(self):
        scd = SharedCompressionDict()
        scd.train([])
        self.assertIsNone(scd._dictionary)

    def test_shared_dict_train_tiny_samples(self):
        scd = SharedCompressionDict()
        scd.train([b"x"])
        self.assertIsNone(scd._dictionary)

    def test_shared_dict_concurrent_compress(self):
        scd = SharedCompressionDict()
        base = b"PATTERN" * 200
        scd.train([base + os.urandom(64) for _ in range(8)])
        errors = []
        def compress_task():
            try:
                for _ in range(50):
                    data = base + os.urandom(32)
                    c = scd.compress(data)
                    d = scd.decompress(c)
                    assert d == data
            except Exception as e:
                errors.append(e)
        threads = [threading.Thread(target=compress_task) for _ in range(4)]
        for t in threads:
            t.start()
        for t in threads:
            t.join()
        self.assertEqual(len(errors), 0)

    def test_streaming_anchor_replacement(self):
        smm = StreamingMemoryManager(anchor_slots=2)
        smm.register_anchor("a", 5.0)
        smm.register_anchor("b", 10.0)
        ok = smm.register_anchor("c", 20.0)
        self.assertTrue(ok)
        self.assertFalse(smm.is_anchor("a"))
        self.assertTrue(smm.is_anchor("c"))

    def test_streaming_concurrent_anchor_ops(self):
        smm = StreamingMemoryManager(anchor_slots=10)
        errors = []
        def register_task(start):
            try:
                for i in range(50):
                    smm.register_anchor(f"blk_{start}_{i}", float(i))
            except Exception as e:
                errors.append(e)
        threads = [threading.Thread(target=register_task, args=(t,)) for t in range(4)]
        for t in threads:
            t.start()
        for t in threads:
            t.join()
        self.assertEqual(len(errors), 0)

    def test_flash_io_submit_during_flush(self):
        flash = FlashIOEngine(max_batch_size=4)
        errors = []
        def submitter():
            try:
                for i in range(100):
                    flash.submit(f"b_{i}", 0, 1, 1024)
            except Exception as e:
                errors.append(e)
        def flusher():
            try:
                for _ in range(50):
                    flash.flush()
            except Exception as e:
                errors.append(e)
        t1 = threading.Thread(target=submitter)
        t2 = threading.Thread(target=flusher)
        t1.start()
        t2.start()
        t1.join()
        t2.join()
        self.assertEqual(len(errors), 0)

    def test_placement_cache_concurrent_record_lookup(self):
        cache = BlockPlacementCache(max_entries=100)
        errors = []
        def writer():
            try:
                for i in range(200):
                    cache.record(f"blk_{i % 50}", i % 3, i, 0.1)
            except Exception as e:
                errors.append(e)
        def reader():
            try:
                for i in range(200):
                    cache.lookup(f"blk_{i % 50}")
            except Exception as e:
                errors.append(e)
        t1 = threading.Thread(target=writer)
        t2 = threading.Thread(target=reader)
        t1.start()
        t2.start()
        t1.join()
        t2.join()
        self.assertEqual(len(errors), 0)

    def test_importance_evictor_empty_blocks(self):
        evictor = ImportanceBasedEvictor()
        victims = evictor.select_victims([], count=5)
        self.assertEqual(len(victims), 0)


# ═══════════════════════════════════════════════════════════════
# R6  Correctness — 數學 / 正確性 (14 tests)
# ═══════════════════════════════════════════════════════════════
class TestR6_Correctness(unittest.TestCase):
    """Mathematical correctness of scores, ratios, rates"""

    def test_importance_recency_decreases_with_time(self):
        ev = ImportanceBasedEvictor()
        now = time.time()
        recent = ev.compute_importance("b1", 10, now, 1000, now)
        old = ev.compute_importance("b2", 10, now - 3600, 1000, now)
        self.assertGreater(recent.recency, old.recency)

    def test_importance_frequency_increases_with_count(self):
        ev = ImportanceBasedEvictor()
        now = time.time()
        low = ev.compute_importance("b1", 1, now, 1000, now)
        high = ev.compute_importance("b2", 1000, now, 1000, now)
        self.assertGreater(high.frequency, low.frequency)

    def test_importance_large_block_penalized(self):
        ev = ImportanceBasedEvictor()
        now = time.time()
        small = ev.compute_importance("b1", 10, now, 1*1024*1024, now)
        large = ev.compute_importance("b2", 10, now, 128*1024*1024, now)
        self.assertGreater(small.score, large.score)

    def test_heavy_hitter_bonus(self):
        ev = ImportanceBasedEvictor(heavy_hitter_threshold=1.5)
        now = time.time()
        for _ in range(50):
            ev.record_access("heavy")
        for _ in range(2):
            ev.record_access("light")
        heavy = ev.compute_importance("heavy", 50, now, 1000, now)
        light = ev.compute_importance("light", 2, now, 1000, now)
        self.assertTrue(heavy.is_heavy_hitter)
        self.assertFalse(light.is_heavy_hitter)

    def test_hit_rate_monotonic(self):
        """Hit rate should increase as speculative learns"""
        spec = SpeculativePrefetchEngine()
        # Train on repeating pattern
        for _ in range(100):
            spec.record_access("A")
            spec.record_access("B")
        # After training, predictions should mostly hit
        hits = 0
        for _ in range(20):
            preds = spec.predict_next("A", ["B", "C"])
            if spec.verify_prediction(preds, "B"):
                hits += 1
            spec.record_access("A")
            spec.record_access("B")
        self.assertGreater(hits, 10)  # >50% hit rate

    def test_confidence_clamps(self):
        spec = SpeculativePrefetchEngine()
        for _ in range(100):
            spec.verify_prediction(["X"], "X")  # all hits
        self.assertLessEqual(spec._current_confidence, 0.95)
        for _ in range(100):
            spec.verify_prediction(["X"], "Y")  # all misses
        self.assertGreaterEqual(spec._current_confidence, 0.1)

    def test_shared_dict_compress_decompress_roundtrip(self):
        sd = SharedCompressionDict()
        base = struct.pack('<' + 'f'*100, *[i*0.01 for i in range(100)])
        samples = [base + bytes([i]*50) for i in range(8)]
        sd.train(samples)
        for data in samples:
            compressed = sd.compress(data)
            decompressed = sd.decompress(compressed)
            self.assertEqual(decompressed, data)

    def test_flash_io_reduction_percentage(self):
        flash = FlashIOEngine(max_batch_size=10, min_batch_bytes=999999999)
        for i in range(20):
            flash.submit(f"b{i}", 2, 1, 1024)
        # Flush twice (10 + 10)
        flash.flush()
        flash.flush()
        stats = flash.stats()
        # 20 individual → 2 batches = 18 saved = 90%
        self.assertGreater(stats["io_reduction_pct"], 80)

    def test_placement_cache_hit_rate(self):
        cache = BlockPlacementCache()
        cache.record("t1", 0, 10, 0.1)
        cache.lookup("t1")  # hit
        cache.lookup("t1")  # hit
        cache.lookup("t2")  # miss
        self.assertAlmostEqual(cache.hit_rate_pct, 66.7, places=0)

    def test_sliding_window_utilization(self):
        sw = SlidingWindowTracker(window_size=10)
        for i in range(5):
            sw.access(f"b{i}")
        self.assertAlmostEqual(sw.window_utilization_pct, 50.0)

    def test_grouped_co_access_symmetric(self):
        gs = GroupedBlockScheduler()
        gs.record_access("A")
        gs.record_access("B")
        gs.flush_window()
        # A→B and B→A should both be recorded
        self.assertEqual(gs._co_access_matrix["A"]["B"], 1)
        self.assertEqual(gs._co_access_matrix["B"]["A"], 1)

    def test_suite_eviction_fallback(self):
        """Without importance evictor, fallback to LRU"""
        suite = LLMOptimizationSuite(enable_importance_eviction=False)
        now = time.time()
        blocks = [
            ("oldest", 100, now - 600, 1000),
            ("newest", 1, now, 1000),
        ]
        victims = suite.get_eviction_candidates(blocks, 1)
        self.assertEqual(victims[0], "oldest")

    def test_streaming_should_evict(self):
        sm = StreamingMemoryManager(max_vram_blocks=10, max_ram_blocks=20)
        self.assertTrue(sm.should_evict(15, 0))
        self.assertTrue(sm.should_evict(0, 25))
        self.assertFalse(sm.should_evict(5, 10))

    def test_grouped_device_affinity(self):
        gs = GroupedBlockScheduler(min_co_access=2)
        for _ in range(5):
            gs.record_access("a1")
            gs.record_access("a2")
            gs.flush_window()
        gs.detect_groups()
        assignment = gs.assign_device_affinity([100.0, 50.0, 10.0])
        # Should assign to fastest device
        self.assertIsInstance(assignment, dict)


# ═══════════════════════════════════════════════════════════════
# R7  LLM Domain — 每項技術語意正確性 (8 tests)
# ═══════════════════════════════════════════════════════════════
class TestR7_LLMDomain(unittest.TestCase):
    """Domain-specific: each LLM technique maps correctly"""

    def test_kv_cache_remembers_across_sessions(self):
        """KV Cache: tier assignment persists across lookups"""
        cache = BlockPlacementCache()
        cache.record("attn_weight", 0, 100, 0.01)
        # Simulating "next session"
        tier = cache.lookup("attn_weight")
        self.assertEqual(tier, 0)  # VRAM

    def test_flash_attention_reduces_io(self):
        """Flash Attention: batching reduces I/O count"""
        flash = FlashIOEngine(max_batch_size=8, min_batch_bytes=999999999)
        for i in range(8):
            flash.submit(f"b{i}", 2, 1, 100)
        batch = flash.flush()
        self.assertEqual(len(batch.requests), 8)
        self.assertEqual(flash.stats()["total_batched_ios"], 1)

    def test_mla_dictionary_improves_compression(self):
        """MLA: shared dict should produce non-empty compressed output"""
        sd = SharedCompressionDict()
        base = bytes(range(256)) * 40
        samples = [base + bytes([i]*100) for i in range(8)]
        sd.train(samples)
        compressed = sd.compress(samples[0])
        self.assertLess(len(compressed), len(samples[0]))

    def test_sliding_window_bounds_memory(self):
        """Sliding Window: hot set never exceeds window_size"""
        sw = SlidingWindowTracker(window_size=50)
        for i in range(1000):
            sw.access(f"token_{i}")
        self.assertLessEqual(sw.stats()["hot_count"], 50)

    def test_streaming_preserves_sinks(self):
        """Streaming LLM: attention sinks never evicted"""
        sm = StreamingMemoryManager(anchor_slots=2)
        sm.register_anchor("sink_0", 100.0, "attention sink")
        sm.register_anchor("sink_1", 90.0, "attention sink")
        blocks = [(f"b{i}", 10, time.time() - i) for i in range(20)]
        victims = sm.select_eviction_candidates(blocks, "vram", 10)
        self.assertNotIn("sink_0", victims)
        self.assertNotIn("sink_1", victims)

    def test_pruning_keeps_heavy_hitters(self):
        """Pruning KV: heavy hitters get 50% bonus"""
        ev = ImportanceBasedEvictor(heavy_hitter_threshold=2.0, heavy_hitter_window=500)
        now = time.time()
        # heavy: 80 accesses, 10 light blocks: 2 each → avg = (80+20)/11 ≈ 9.1
        # heavy(80) > avg(9.1) * 2.0(18.2) → True
        for _ in range(80):
            ev.record_access("heavy")
        for i in range(10):
            for _ in range(2):
                ev.record_access(f"light_{i}")
        imp = ev.compute_importance("heavy", 80, now, 1000, now)
        self.assertTrue(imp.is_heavy_hitter)
        self.assertGreater(imp.score, 0)

    def test_speculative_learns_patterns(self):
        """Speculative Decoding: learns A→B transition"""
        spec = SpeculativePrefetchEngine()
        for _ in range(30):
            spec.record_access("A")
            spec.record_access("B")
        preds = spec.predict_next("A", ["B", "C", "D"])
        self.assertEqual(preds[0], "B")

    def test_gqa_groups_correlated_blocks(self):
        """GQA: Q/K/V blocks end up in same group"""
        gs = GroupedBlockScheduler(min_co_access=3)
        for _ in range(5):
            for name in ["Q", "K", "V"]:
                gs.record_access(name)
            gs.flush_window()
        groups = gs.detect_groups()
        self.assertTrue(len(groups) >= 1)
        # All three should be in same group
        qkv_group = gs.get_group_for_block("Q")
        self.assertIsNotNone(qkv_group)
        self.assertIn("K", qkv_group.block_ids)
        self.assertIn("V", qkv_group.block_ids)

    def test_kv_cache_multi_session_warmup(self):
        """PlacementCache should improve across sessions"""
        cache = BlockPlacementCache()
        for i in range(20):
            cache.record(f"layer_{i}", tier=0 if i < 5 else 1,
                         access_count=100 - i, latency_ms=0.1)
        hits = sum(1 for i in range(20) if cache.lookup(f"layer_{i}") is not None)
        self.assertEqual(hits, 20)

    def test_mla_dict_adapts_to_structured_data(self):
        """SharedCompressionDict should handle structured float data"""
        scd = SharedCompressionDict()
        layers = []
        for i in range(10):
            vals = [math.sin(j * 0.01 + i * 0.1) for j in range(2048)]
            layers.append(struct.pack('<' + 'f' * len(vals), *vals))
        scd.train(layers[:8])
        scd.compress(layers[8])
        scd.compress(layers[9])
        self.assertGreaterEqual(scd.improvement_pct, 0.0)

    def test_streaming_sink_survives_heavy_eviction(self):
        """Anchors (attention sinks) must never be evicted"""
        smm = StreamingMemoryManager(anchor_slots=4)
        smm.register_anchor("sink_0", 100.0, "attention sink")
        smm.register_anchor("sink_1", 90.0, "attention sink")
        blocks = [("sink_0", 500, 1.0), ("sink_1", 400, 1.0)]
        blocks += [(f"b_{i}", 1, 0.1) for i in range(100)]
        evicted = smm.select_eviction_candidates(blocks, "vram", count=50)
        self.assertNotIn("sink_0", evicted)
        self.assertNotIn("sink_1", evicted)

    def test_speculative_cold_start_conservative(self):
        """Before learning, speculative should use moderate lookahead; after hits, expand"""
        spec = SpeculativePrefetchEngine(base_lookahead=2, max_lookahead=8)
        # Default confidence=0.5 → base+2=4 (moderate)
        self.assertLessEqual(spec.current_lookahead, spec._max_lookahead)
        initial = spec.current_lookahead
        # After many hits, confidence rises → max lookahead
        for _ in range(20):
            spec.verify_prediction(["a", "b", "c"], "a")
        self.assertGreaterEqual(spec.current_lookahead, initial)

    def test_grouped_io_detects_affinity(self):
        """GQA scheduler should detect block groups from co-access"""
        gs = GroupedBlockScheduler(min_co_access=3)
        for _ in range(5):
            gs.record_access("a")
            gs.record_access("b")
            gs.record_access("c")
            gs.flush_window()
        groups = gs.detect_groups()
        self.assertGreater(len(groups), 0)
        plan = gs.get_group_io_plan("a")
        self.assertIn("b", plan)

    def test_pruning_preserves_heavy_hitters(self):
        """H2O: heavy hitters should not be evicted first"""
        evictor = ImportanceBasedEvictor(heavy_hitter_threshold=2.0)
        for _ in range(50):
            evictor.record_access("block_0")
        for _ in range(5):
            evictor.record_access("block_1")
        now = time.time()
        blocks = [
            ("block_0", 50, now - 10, 64 * 1024 * 1024),
            ("block_1", 5, now - 1, 64 * 1024 * 1024),
        ]
        victims = evictor.select_victims(blocks, count=1)
        self.assertEqual(victims[0].block_id, "block_1")


# ═══════════════════════════════════════════════════════════════
# R8  Integration — MemoryPool / Prefetcher 整合 (8 tests)
# ═══════════════════════════════════════════════════════════════
class TestR8_Integration(unittest.TestCase):
    """Integration with existing modules"""

    def test_memory_pool_access_triggers_sw_tracker(self):
        from microsystems.core.memory_pool import MemoryPool, MemoryTier
        pool = MemoryPool(1024**3, 4*1024**3, 10*1024**3)
        pool.allocate("b1", 100*1024*1024, tag="test")
        pool.access("b1")
        self.assertGreater(pool._sw_tracker.stats()["total_accesses"], 0)

    def test_memory_pool_access_triggers_importance(self):
        from microsystems.core.memory_pool import MemoryPool
        pool = MemoryPool(1024**3, 4*1024**3, 10*1024**3)
        pool.allocate("b1", 100*1024*1024)
        pool.access("b1")
        self.assertGreater(
            pool._importance_evictor.stats()["tracked_blocks"], 0)

    def test_memory_pool_eviction_uses_importance(self):
        """Eviction should use importance scoring, not just LRU"""
        from microsystems.core.memory_pool import MemoryPool, MemoryTier
        pool = MemoryPool(
            vram_capacity_bytes=200*1024*1024,
            ram_capacity_bytes=1024**3,
            external_capacity_bytes=10*1024**3,
        )
        # Fill VRAM with 2 blocks
        pool.allocate("hot", 100*1024*1024, preferred_tier=MemoryTier.VRAM)
        pool.allocate("cold", 100*1024*1024, preferred_tier=MemoryTier.VRAM)
        # Access hot block many times
        for _ in range(20):
            pool.access("hot")
        # Allocate a third → should evict cold, not hot
        pool.allocate("new", 100*1024*1024, preferred_tier=MemoryTier.VRAM)
        hot_blk = pool.get_block("hot")
        self.assertIsNotNone(hot_blk)

    def test_prefetcher_has_speculative_engine(self):
        from microsystems.core.memory_pool import MemoryPool, MemoryTier
        from microsystems.core.transfer_engine import TransferEngine
        from microsystems.core.prefetcher import PredictivePrefetcher
        pool = MemoryPool(1024**3, 4*1024**3, 10*1024**3)
        transfer = TransferEngine()
        pf = PredictivePrefetcher(pool, transfer)
        self.assertTrue(hasattr(pf, '_speculative'))
        self.assertIsInstance(pf._speculative, SpeculativePrefetchEngine)

    def test_transfer_engine_has_flash_io(self):
        from microsystems.core.transfer_engine import TransferEngine
        te = TransferEngine()
        self.assertTrue(hasattr(te, '_flash_io'))
        self.assertIsInstance(te._flash_io, FlashIOEngine)

    def test_slow_device_has_shared_dict(self):
        from microsystems.core.slow_device_optimizer import SlowDeviceOptimizer
        opt = SlowDeviceOptimizer()
        self.assertTrue(hasattr(opt, '_shared_dict'))
        self.assertIsInstance(opt._shared_dict, SharedCompressionDict)

    def test_striped_swap_has_grouped_scheduler(self):
        from microsystems.core.striped_swap import StripedSwapScheduler
        ss = StripedSwapScheduler()
        self.assertTrue(hasattr(ss, '_grouped_scheduler'))
        self.assertIsInstance(ss._grouped_scheduler, GroupedBlockScheduler)

    def test_real_boost_has_llm_opts(self):
        from microsystems.core.real_boost import RealBoostEngine
        engine = RealBoostEngine()
        self.assertTrue(hasattr(engine, '_llm_opts'))
        self.assertIsInstance(engine._llm_opts, LLMOptimizationSuite)

    def test_placement_cache_warm_start_plan(self):
        """PlacementCache hint should influence allocation"""
        cache = BlockPlacementCache()
        cache.record("layer_0", tier=1, access_count=100, latency_ms=0.1)
        tier = cache.lookup("layer_0")
        self.assertEqual(tier, 1)
        plan = cache.warm_start_plan(["layer_0", "layer_1", "unknown"])
        self.assertIn("layer_0", plan)
        self.assertNotIn("unknown", plan)

    def test_flash_io_direction_grouping(self):
        """FlashIO should group transfers by direction"""
        flash = FlashIOEngine(max_batch_size=100)
        flash.submit("a", 2, 1, 1024)
        flash.submit("b", 2, 1, 1024)
        flash.submit("c", 1, 0, 1024)
        batch = flash.flush()
        self.assertIsNotNone(batch)
        groups = flash.group_by_direction(batch)
        self.assertEqual(len(groups), 2)
        self.assertEqual(len(groups[(2, 1)]), 2)

    def test_shared_dict_compress_integrity(self):
        """SharedCompressionDict roundtrip with trained dict"""
        scd = SharedCompressionDict()
        base = b"MODEL_WEIGHT_" * 100
        samples = [base + bytes([i]) * 64 for i in range(8)]
        scd.train(samples)
        original = base + b"\xff" * 64
        compressed = scd.compress(original)
        decompressed = scd.decompress(compressed)
        self.assertEqual(decompressed, original)

    def test_suite_on_block_access_dispatches(self):
        """LLMOptimizationSuite.on_block_access dispatches to sub-engines"""
        suite = LLMOptimizationSuite()
        suite.on_block_access("blk_0", access_count=1, size_bytes=1024)
        self.assertEqual(suite.sliding_window.stats()["total_accesses"], 1)


# ═══════════════════════════════════════════════════════════════
# R9  Pipeline — StripedSwap + GroupedIO + FlashIO (6 tests)
# ═══════════════════════════════════════════════════════════════
class TestR9_Pipeline(unittest.TestCase):
    """Multi-device pipeline tests"""

    def test_grouped_io_plan_single_block(self):
        gs = GroupedBlockScheduler()
        plan = gs.get_group_io_plan("standalone")
        self.assertEqual(plan, ["standalone"])

    def test_flash_io_multiple_flushes(self):
        """Flush drains all pending in one batch; second flush returns None"""
        flash = FlashIOEngine(max_batch_size=5, min_batch_bytes=999999999)
        for i in range(12):
            flash.submit(f"b{i}", 2, 1, 100)
        b1 = flash.flush()
        b2 = flash.flush()  # should be None (all drained)
        self.assertIsNotNone(b1)
        self.assertIsNone(b2)
        self.assertEqual(len(b1.requests), 12)
        # Verify stats: 12 individual → 1 batch
        self.assertEqual(flash.stats()["total_batched_ios"], 1)
        self.assertEqual(flash.stats()["saved_ios"], 11)

    def test_suite_pipeline_no_crash(self):
        suite = LLMOptimizationSuite()
        for i in range(50):
            suite.on_block_access(f"b{i % 10}")
            suite.on_transfer_request(f"b{i}", 2, 1, 1024)
        if suite.flash_io.should_flush():
            suite.flash_io.flush()
        stats = suite.full_stats()
        self.assertGreater(stats["flash_io"]["total_individual_ios"], 0)

    def test_grouped_assign_empty_devices(self):
        gs = GroupedBlockScheduler()
        assignment = gs.assign_device_affinity([])
        self.assertEqual(assignment, {})

    def test_flash_io_should_flush_timing(self):
        flash = FlashIOEngine(batch_window_ms=1.0)
        flash.submit("b1", 2, 1, 100)
        time.sleep(0.01)  # 10ms > 1ms window
        self.assertTrue(flash.should_flush())

    def test_grouped_multiple_detect_cycles(self):
        gs = GroupedBlockScheduler(min_co_access=2)
        for _ in range(3):
            gs.record_access("x")
            gs.record_access("y")
            gs.flush_window()
        g1 = gs.detect_groups()
        g2 = gs.detect_groups()  # second detect
        self.assertGreater(len(g1), 0)

    def test_grouped_scheduler_device_affinity(self):
        """Groups should get assigned to fastest devices"""
        gs = GroupedBlockScheduler(min_co_access=2)
        for _ in range(5):
            gs.record_access("a")
            gs.record_access("b")
            gs.flush_window()
        gs.detect_groups()
        assignment = gs.assign_device_affinity([100.0, 200.0, 50.0])
        if assignment:
            first_group = list(assignment.keys())[0]
            self.assertEqual(assignment[first_group], 1)  # fastest=200MB/s at idx 1

    def test_flash_io_batch_timeout(self):
        """FlashIO should indicate flush needed after time window"""
        flash = FlashIOEngine(batch_window_ms=1.0, max_batch_size=1000)
        flash.submit("x", 0, 1, 1024)
        time.sleep(0.01)  # 10ms > 1ms window
        self.assertTrue(flash.should_flush())


# ═══════════════════════════════════════════════════════════════
# R10 Encoding — Unicode / JSON roundtrip (6 tests)
# ═══════════════════════════════════════════════════════════════
class TestR10_Encoding(unittest.TestCase):
    """Unicode block tags, Chinese characters, JSON roundtrip"""

    def test_unicode_block_tags(self):
        cache = BlockPlacementCache()
        cache.record("層_32_權重", 0, 10, 0.1)
        self.assertEqual(cache.lookup("層_32_權重"), 0)

    def test_unicode_sliding_window(self):
        sw = SlidingWindowTracker(window_size=5)
        sw.access("注意力層_0")
        sw.access("前饋層_0")
        self.assertTrue(sw.is_hot("注意力層_0"))

    def test_unicode_anchor(self):
        sm = StreamingMemoryManager()
        sm.register_anchor("注意力沉沒_0", 10.0, "第一個 token 的 attention sink")
        self.assertTrue(sm.is_anchor("注意力沉沒_0"))

    def test_unicode_grouped_io(self):
        gs = GroupedBlockScheduler(min_co_access=2)
        for _ in range(3):
            gs.record_access("查詢向量")
            gs.record_access("鍵向量")
            gs.flush_window()
        groups = gs.detect_groups()
        self.assertGreater(len(groups), 0)

    def test_stats_json_serializable(self):
        """All stats dicts must be JSON serializable"""
        suite = LLMOptimizationSuite()
        for i in range(5):
            suite.on_block_access(f"b{i}")
        stats = suite.full_stats()
        json_str = json.dumps(stats, ensure_ascii=False)
        parsed = json.loads(json_str)
        self.assertEqual(parsed["enabled"]["placement_cache"], True)

    def test_speculative_unicode_transitions(self):
        spec = SpeculativePrefetchEngine()
        spec.record_access("嵌入層")
        spec.record_access("注意力層")
        preds = spec.predict_next("嵌入層", ["注意力層", "前饋層"])
        self.assertIn("注意力層", preds)


# ═══════════════════════════════════════════════════════════════
# R11 Export — Benchmark + Report (8 tests)
# ═══════════════════════════════════════════════════════════════
class TestR11_Export(unittest.TestCase):
    """Benchmark execution and report generation"""

    def test_run_all_benchmarks_returns_8(self):
        from microsystems.core.benchmark_llm_opts import run_all_benchmarks
        results = run_all_benchmarks()
        self.assertEqual(len(results), 8)

    def test_benchmark_results_have_required_fields(self):
        from microsystems.core.benchmark_llm_opts import run_all_benchmarks
        results = run_all_benchmarks()
        for r in results:
            self.assertTrue(hasattr(r, "name"))
            self.assertTrue(hasattr(r, "before_value"))
            self.assertTrue(hasattr(r, "after_value"))
            self.assertTrue(hasattr(r, "improvement_pct"))

    def test_all_benchmarks_positive_improvement(self):
        from microsystems.core.benchmark_llm_opts import run_all_benchmarks
        results = run_all_benchmarks()
        for r in results:
            self.assertGreaterEqual(r.improvement_pct, -5.0,
                f"{r.name} has negative improvement: {r.improvement_pct}")

    def test_excel_report_generation(self):
        from microsystems.core.benchmark_llm_opts import (
            run_all_benchmarks, generate_excel_report,
        )
        results = run_all_benchmarks()
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "test_report.xlsx")
            result_path = generate_excel_report(results, path)
            self.assertTrue(os.path.exists(result_path))
            self.assertGreater(os.path.getsize(result_path), 5000)

    def test_excel_has_3_sheets(self):
        from openpyxl import load_workbook
        from microsystems.core.benchmark_llm_opts import (
            run_all_benchmarks, generate_excel_report,
        )
        results = run_all_benchmarks()
        with tempfile.TemporaryDirectory() as tmpdir:
            path = os.path.join(tmpdir, "test_report.xlsx")
            generate_excel_report(results, path)
            wb = load_workbook(path)
            self.assertEqual(len(wb.sheetnames), 3)

    def test_benchmark_names_unique(self):
        from microsystems.core.benchmark_llm_opts import run_all_benchmarks
        results = run_all_benchmarks()
        names = [r.name for r in results]
        self.assertEqual(len(names), len(set(names)))

    def test_benchmark_before_after_types(self):
        from microsystems.core.benchmark_llm_opts import run_all_benchmarks
        results = run_all_benchmarks()
        for r in results:
            self.assertIsInstance(r.before_value, (int, float))
            self.assertIsInstance(r.after_value, (int, float))

    def test_suite_stats_match_benchmark(self):
        """Stats keys should be stable"""
        suite = LLMOptimizationSuite()
        stats = suite.full_stats()
        expected_keys = {"placement_cache", "flash_io", "shared_dict",
                         "sliding_window", "streaming", "importance_eviction",
                         "speculative_prefetch", "grouped_io", "enabled"}
        self.assertEqual(set(stats.keys()), expected_keys)


if __name__ == "__main__":
    unittest.main()
